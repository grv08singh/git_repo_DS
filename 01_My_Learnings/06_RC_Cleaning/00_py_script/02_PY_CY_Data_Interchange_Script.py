import os
import openpyxl
import pandas


def empty_py(wb_py):
    #deleting data from RC1 PY
    ws_py1 = wb_py["RC-1(PY)"]
    max_row = ws_py1.max_row
    ws_py1.delete_rows(1, max_row)
    
    #deleting data from RC2 PY
    ws_py2 = wb_py["RC-2(PY)"]
    max_row = ws_py2.max_row
    ws_py2.delete_rows(1, max_row)
    
    #deleting data from RC3 PY
    ws_py3 = wb_py["RC-3(PY)"]
    max_row = ws_py3.max_row
    ws_py3.delete_rows(1, max_row)
    
    #deleting data from RC7 PY
    ws_py7 = wb_py["RC-7(PY)"]
    max_row = ws_py7.max_row
    ws_py7.delete_rows(1, max_row)
    
    return wb_py


def fill_py(wb_py, wb_cy):
    
    #copying RC1 CY to PY
    ws_py1 = wb_py["RC-1(PY)"]
    ws_cy1 = wb_cy["RC-1"]
    for row in ws_cy1.iter_rows():
        for cell in row:
            ws_py1[cell.coordinate].value = cell.value
    
    #copying RC2 CY to PY
    ws_py2 = wb_py["RC-2(PY)"]
    ws_cy2 = wb_cy["RC-2"]
    for row in ws_cy2.iter_rows():
        for cell in row:
            ws_py2[cell.coordinate].value = cell.value
    
    #copying RC3 CY to PY
    ws_py3 = wb_py["RC-3(PY)"]
    ws_cy3 = wb_cy["RC-3"]
    for row in ws_cy3.iter_rows():
        for cell in row:
            ws_py3[cell.coordinate].value = cell.value
    
    #copying RC7 CY to PY
    ws_py7 = wb_py["RC-7(PY)"]
    ws_cy7 = wb_cy["RC-7"]
    for row in ws_cy7.iter_rows():
        for cell in row:
            ws_py7[cell.coordinate].value = cell.value
    
    return wb_py


def empty_cy(wb_cy):

    #deleting data from RC1 CY
    ws_cy1 = wb_cy["RC-1"]
    max_row = ws_cy1.max_row
    ws_cy1.delete_rows(1, max_row)
    
    #deleting data from RC2 CY
    ws_cy2 = wb_cy["RC-2"]
    max_row = ws_cy2.max_row
    ws_cy2.delete_rows(1, max_row)
    
    #deleting data from RC3 CY
    ws_cy3 = wb_cy["RC-3"]
    max_row = ws_cy3.max_row
    ws_cy3.delete_rows(1, max_row)
    
    #deleting data from RC7 CY
    ws_cy7 = wb_cy["RC-7"]
    max_row = ws_cy7.max_row
    ws_cy7.delete_rows(1, max_row)
    
    return wb_cy


def fill_cy(wb_cy, wb_rc1237):
    
    #copying newly created RC1 to CY
    ws_cy1 = wb_cy["RC-1"]
    ws_rc1 = wb_rc1237["RC1"]
    for row in ws_rc1.iter_rows():
        for cell in row:
            ws_cy1[cell.coordinate].value = cell.value

    #copying newly created RC2 to CY
    ws_cy2 = wb_cy["RC-2"]
    ws_rc2 = wb_rc1237["RC2"]
    for row in ws_rc2.iter_rows():
        for cell in row:
            ws_cy2[cell.coordinate].value = cell.value

    #copying newly created RC3 to CY
    ws_cy3 = wb_cy["RC-3"]
    ws_rc3 = wb_rc1237["RC3"]
    for row in ws_rc3.iter_rows():
        for cell in row:
            ws_cy3[cell.coordinate].value = cell.value

    #copying newly created RC7 to CY
    ws_cy7 = wb_cy["RC-7"]
    ws_rc7 = wb_rc1237["RC7"]
    for row in ws_rc7.iter_rows():
        for cell in row:
            ws_cy7[cell.coordinate].value = cell.value
    
    return wb_cy




if __name__ == "__main__":
    
    #loading newly created RCs
    os.chdir('..')
    os.chdir('04_Clean_RCs')
    wb_rc1237 = openpyxl.load_workbook("RC1237_clean.xlsx")
    ws_rc1 = wb_rc1237["RC1"]
    ws_rc2 = wb_rc1237["RC2"]
    ws_rc3 = wb_rc1237["RC3"]
    ws_rc7 = wb_rc1237["RC7"]

    #read previous year's CY & PY
    os.chdir('..')
    os.chdir('03_PY_CY_Old')
    wb_py = openpyxl.load_workbook("PY State Details & Capital-N .xlsx")
    wb_cy = openpyxl.load_workbook("CY State Details & Capital-N.xlsx")

    #updating CY, PY files
    wb_py = empty_py(wb_py)
    wb_py = fill_py(wb_py, wb_cy)
    wb_cy = empty_cy(wb_cy)
    wb_cy = fill_cy(wb_cy, wb_rc1237)

    #saving new PY, CY excel files in folder "06_PY_CY_New"
    os.chdir('..')
    os.chdir('06_PY_CY_New')
    wb_py.save("PY State Details & Capital-N .xlsx")
    wb_cy.save("CY State Details & Capital-N.xlsx")