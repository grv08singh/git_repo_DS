#%%
from pathlib import Path
import numpy as np
import pandas as pd
import pyodbc
import sys
import os
import time
from datetime import date
import re
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings as wr
wr.filterwarnings('ignore')
#%%

#%%
def clean_rc123(details_1p0):
    
    #get data from the table - RC1
    year = details_1p0['year']
    state_cd = details_1p0['state_cd']
    season_cd = details_1p0['season_cd']
    season_nm = 'kharif' if season_cd in [1,2,3] else 'rabi' if season_cd in [4,5] else 'Season Incorrect'
    st_cd = details_1p0['st_cd']
    novp = details_1p0['novp']
    novcod = details_1p0['novcod']
    ddg = details_1p0['ddg']
    ddtrs = details_1p0['ddtrs']
    crop_set = details_1p0['crop_set']
    crop_set.update([0,99])
    crops_rc2 = sorted(list(crop_set))
    
    crops_rc3 = crops_rc2.copy()
    crops_rc3.remove(0)

    
    #set database file path
    folder_path = Path('./01_software')
    for file_path in folder_path.iterdir():
        if file_path.is_file():
            if season_nm.lower() in file_path.name.lower():
                accdb_path = "./" + "/".join(str(file_path).split("\\"))
    
    #create connection with mdb database
    conn_str = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        f'DBQ={accdb_path};'
    )
    conn = pyodbc.connect(conn_str)
    
    cols1 = ['RC', 'ST', 'YR', 'SESON', 'NOVP', 'NOVCOD', 'STAT', 'DIST', 'STRA',
           'VILL', 'EPC', 'NOVPC', 'NOVTRS', 'CADC', 'TLC', 'MAC', 'MUC', 'DDG',
           'SGC', 'GCC', 'LFOG', 'ROG', 'DDTRS', 'TRSSC', 'TRSSF', 'DCKC', 'HSSN',
           'GEOA', 'TCROPA', 'TNSSN', 'TGEOASN', 'RC61', 'RC62', 'RC63', 'REJ']
    
    sql_query1 = 'SELECT * FROM "Copy Of RC1"'
    df1 = pd.read_sql(sql_query1, conn)[cols1]
    
    df1 = (df1
        .drop_duplicates()
        .apply(pd.to_numeric, errors='coerce')
        .query("STAT == @state_cd and SESON == @season_cd and ST == @st_cd")
        .assign(
            REJ = lambda x : np.where(x['REJ']==1, 1, 0)
            ,NOVPC1 = lambda x : x['NOVPC'].fillna(1)
            ,NOVPC = lambda x : np.where(x['NOVPC1']==0, 1, x['NOVPC1'])
            ,NOVTRS1 = lambda x : x['NOVTRS'].fillna(1)
            ,NOVTRS = lambda x : np.where(x['NOVTRS1']==0, 1, x['NOVTRS1'])
            ,CADC = lambda x : np.where(np.isin(x['CADC'], [1,2,3,4]), x['CADC'], 9)
            ,TLC = lambda x : np.where(x['CADC']==3, 0, np.where(np.isin(x['CADC'], [4,9]), 8, x['TLC']))
            ,MAC = lambda x : np.where(x['MAC']==1, 1, 0)
            ,MUC = lambda x : np.where(x['MAC']==1, x['MUC'], None)
            ,TRSSC = lambda x : np.where(x['TRSSC']==3, 8, x['TRSSC'])
            ,NOVP = novp
            ,DDG = ddg
            ,DDTRS = ddtrs
            ,TCROPA = lambda x : x['TCROPA'].fillna(0)
            ,GEOA = lambda x : x['GEOA'].fillna(0)
            ,YR = year+1
            ,NOVCOD = novcod
            ,LFOG = lambda x : np.where(~np.isin(x['LFOG'], [0,1]), None, x['LFOG'])
            ,ROG = lambda x : np.where(~np.isin(x['ROG'], [0,1]), None, x['ROG'])
            ,TRSSF = lambda x : np.where(x['TRSSF'].isna(), None, x['TRSSF'])
        )
        .drop(columns = ['NOVPC1','NOVTRS1'])
    )
    
    non_rej_rc1 = df1[df1['REJ']!=1]
    rej_rc1 = df1[df1['REJ']==1].assign(
    	EPC = None
    	,NOVPC = None
    	,NOVTRS = None
    	,CADC = None
    	,TLC = None
    	,MAC = None
    	,MUC = None
    	,DDG = None
    	,SGC = None
    	,GCC = None
    	,LFOG = None
    	,ROG = None
    	,DDTRS = None
    	,TRSSC = None
    	,TRSSF = None
    	,DCKC = None
    	,HSSN = None
    	,GEOA = None
    	,TCROPA = None
    	,TNSSN = None
    	,TGEOASN = None
    	,RC61 = None
    	,RC62 = None
    	,RC63 = None
    )
    rc1 = pd.concat([non_rej_rc1, rej_rc1], ignore_index=True)
    
    
    
    
    #get data from the table - RC2
    cols2 = ['RC', 'ST', 'YR', 'SESON', 'STAT', 'DIST', 'STRA', 'VILL', 'EPC', 
             'HSSN', 'TNSSN', 'CROP', 'VARC', 'ARSU', 'ARSI', 'ARPU', 'ARPI']
    
    sql_query2 = 'SELECT * FROM "Copy Of RC2"'
    df2 = pd.read_sql(sql_query2, conn)[cols2]
    df2 = (df2
        .drop_duplicates()
        .apply(pd.to_numeric, errors='coerce')
        .query("STAT == @state_cd and SESON == @season_cd and ST == @st_cd")
        .assign(
            YR = year+1
            ,VARC = lambda x : np.where(x['CROP']==0, 0, x['VARC'])
            ,CROP = lambda x : np.where(x['VARC']==0, 0, np.where(np.isin(x['CROP'], crops_rc2), x['CROP'], 99))
        )
    )
    df2['CROP'].isin(crops_rc2)
    rc2 = df2.copy()
    
    
    
    
    
    #get data from the table - RC3
    cols3 = ['RC', 'ST', 'YR', 'SESON', 'STAT', 'DIST', 'STRA',
             'VILL', 'SN', 'CROP', 'VARC', 'IRRC', 'ERC']

    
    sql_query3 = 'SELECT * FROM "Copy Of RC3"'
    df3 = pd.read_sql(sql_query3, conn)[cols3]
    df3 = (df3
        .apply(pd.to_numeric, errors='coerce')
        .query("STAT == @state_cd and SESON == @season_cd and ST == @st_cd")
        .assign(
            YR = year+1
            ,CROP = lambda x : np.where(np.isin(x['CROP'], crops_rc3), x['CROP'], 99)
        )
    )
    
    rc1_bkp = rc1.copy()
    #creating schedule unique id in RC1
    rc1['unique_id'] = rc1['STAT'].astype(str) + "_" + rc1['SESON'].astype(str) + "_" + rc1['ST'].astype(str) + "_" + rc1['STRA'].astype(str) + "_" + rc1['VILL'].astype(str)+ "_" + rc1['DIST'].astype(str)
    
    
    #creating schedule unique id in df3
    df3['unique_id'] = df3['STAT'].astype(str) + "_" + df3['SESON'].astype(str) + "_" + df3['ST'].astype(str) + "_" + df3['STRA'].astype(str) + "_" + df3['VILL'].astype(str)+ "_" + df3['DIST'].astype(str)
    
    #Left Join on df3 with rc1_epc1
    rc1_epc = rc1[['unique_id','EPC']]
    rc3 = pd.merge(df3, rc1_epc, how='left', left_on='unique_id', right_on='unique_id')
    rc3 = rc3[rc3['EPC'] == 1].drop(columns=['unique_id','EPC'])
    
    
    
    
    
    conn.close()
    return rc1_bkp, rc2, rc3
#%%



#%%
def clean_rc7(details_2p0):

    #get data from the table - RC7
    year = details_2p0['year']
    state_cd = details_2p0['state_cd']
    season_cd = details_2p0['season_cd']
    st_cd = details_2p0['st_cd']
    #dataframe containing crop details
    crop_details = details_2p0['crop_details']
    
    #set database file paths
    folder_path = Path('./01_software')
    for file_path in folder_path.iterdir():
        
        if file_path.is_file():
            if 'NSSAS' in file_path.name:
                mdb_path = "./" + "/".join(str(file_path).split("\\"))
    
    #create connection with mdb database
    conn_str = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        f'DBQ={mdb_path};'
    )
    conn = pyodbc.connect(conn_str)
    
    #get data from the table - Sch20RC7
    cols7 = ['RC', 'ST', 'YR', 'SESON', 'CROP', 'NEXP', 'NEXNR', 'CONFAC', 
             'STAT', 'DIST', 'VILL', 'RSV', 'EXPNO', 'STAGIN', 'REXML', 
             'CCE', 'CFC', 'RCSSN', 'E1', 'E2', 'E9', 'E10', 'E11', 'E12', 
             'ADH', 'QTYG', 'QTYP', 'E13', 'EQST', 'EQSB', 'EQSW', 'EQSP', 
             'E14', 'STYPE', 'E3', 'SVARC', 'SEEDR', 'E4', 'IRR', 'E5', 
             'IRRP', 'NIRR', 'FERTC', 'E6', 'QTN', 'CODN', 'QTP', 'CODP', 
             'QTK', 'CODK', 'MANUR', 'E7', 'PEST', 'E8', 'E15', 'REJ', 
             'CROPER', 'CROPCON', 'AFFECT']
    
    sql_query7 = 'SELECT * FROM "Sch20RC7"'
    rc7 = pd.read_sql(sql_query7, conn)[cols7]
    rc7 = (rc7
        .drop_duplicates()
        .apply(pd.to_numeric, errors="coerce")
        .query("STAT == @state_cd and SESON == @season_cd and ST == @st_cd")
        .assign(
            YR = year
            ,CROP = lambda x : x['CROP'].replace({55:12, 57:56, 52:14})       #adjusting cotton & sugarcane crop code
            ,REJ = lambda x : x['REJ'].apply(lambda y : 1 if y==1 else 0)
            ,RSV = lambda x: x['RSV'].where(x["RSV"].isin([0, 1, 2, 3, 9]), 0)
            ,REXML = lambda x: x["REXML"].where(
                        ~(x["STAGIN"].isin([2, 3]) & ~x["REXML"].isin([1, 2, 3, 4, 9])),
                        9
            )
            ,CFC = lambda x: x["CFC"].where(
                        ~(x["STAGIN"].isin([1, 2]) & ~x["CFC"].isin([1, 2, 3, 4])),
                        4
            )
            ,RCSSN = lambda x: x["RCSSN"].where(
                ~(x["STAGIN"].isin([1, 2]) & ~x["RCSSN"].isin([0, 1, 2, 3, 9])),
                0
            )
            ,EQST = lambda x: x["EQST"].where(
                ~(x["STAGIN"].isin([1, 2]) & x["EQST"].isna()),
                9
            )
            ,EQSB = lambda x: x["EQSB"].where(
                ~(x["STAGIN"].isin([1, 2]) & x["EQSB"].isna()),
                9
            )
            ,EQSW = lambda x: x["EQSW"].where(
                ~(x["STAGIN"].isin([1, 2]) & x["EQSW"].isna()),
                9
            )
            ,EQSP = lambda x: x["EQSP"].where(
                ~(x["STAGIN"].isin([1, 2]) & x["EQSP"].isna()),
                9
            )
            ,QTYG = lambda x: (
                x["QTYG"]
                .where(x["STAGIN"] != 2, None)
                .where(~(x["STAGIN"].eq(1) & x["CROPCON"].eq(4)), 0)
            )
            ,QTYP = lambda x: (
                x["QTYP"]
                .where(x["STAGIN"] != 2, None)
                .where(~(x["STAGIN"].eq(1) & x["CROPCON"].eq(4)), 0)
            )
            ,E13 = lambda x : np.where((x['REJ']==1) | np.isin(x["STAGIN"], [2,3]), None, x["E13"])
            ,E14 = lambda x : np.where((x['REJ']==1) | (x['STAGIN']==3), None, np.where(np.isin(x['CROP'], [14,33,44]), 8, x['E14']))
            ,AFFECT = lambda x: x['AFFECT'].where(~((x['STAGIN'].isin([1,2])) & (x['CROPCON'].isin([1,2]))), 0)
            ,NIRR = lambda x : np.where(x['NIRR'].isna(), 0, x['NIRR'])
            ,IRR = lambda x : np.where(x['NIRR']>0, 1, 0)
            ,CODN = lambda x : np.where(x['QTN']>0, 1, x['CODN'])
            ,CODP = lambda x : np.where(x['QTP']>0, 2, x['CODP'])
            ,CODK = lambda x : np.where(x['QTK']>0, 3, x['CODK'])
            ,QTN = lambda x : np.where(x['CODN']==9, 0, np.where(x['CODN']==8, None, x['QTN']))
            ,QTP = lambda x : np.where(x['CODP']==9, 0, np.where(x['CODP']==8, None, x['QTP']))
            ,QTK = lambda x : np.where(x['CODK']==9, 0, np.where(x['CODK']==8, None, x['QTK']))
            ,FERTC = lambda x : np.where(((x['QTN']>0) | (x['QTP']>0) | (x['QTK']>0) | (x['CODN']==9) | (x['CODP']==9) | (x['CODK']==9)), 1, 0)
        )
    )
    
    lost_cols = ['CCE', 'CFC', 'RCSSN', 'E1', 'E2', 'E9', 'E10', 'E11', 'E12', 
                 'ADH', 'QTYG', 'QTYP', 'E13', 'EQST', 'EQSB', 'EQSW', 'EQSP',
                 'E14', 'STYPE', 'E3', 'SVARC', 'SEEDR', 'E4', 'IRR', 'E5',
                 'IRRP', 'NIRR', 'FERTC', 'E6', 'QTN', 'CODN', 'QTP', 'CODP',
                 'QTK', 'CODK', 'MANUR', 'E7', 'PEST', 'E8', 'E15','CROPER', 
                 'CROPCON', 'AFFECT']
    
    for col in lost_cols:
        rc7.loc[rc7['STAGIN']==3, col] = np.nan

    for i in range(crop_details.shape[0]):
        crop = crop_details.loc[i,'crop_cd']
        plan = crop_details.loc[i,'nexp']
        ccf = crop_details.loc[i,'ccf']
        
        rc7.loc[((rc7['ST']==st_cd) & (rc7['CROP']==crop)), 'NEXP'] = plan
        
        received = rc7[(rc7['ST']==st_cd) & (rc7['CROP']==crop)].shape[0]
        rc7.loc[((rc7['ST']==st_cd) & (rc7['CROP']==crop)), 'NEXNR'] = rc7['NEXP'].sub(received, fill_value=0)

        if ccf < 1:
            ccf = np.floor(ccf*10000)
        rc7.loc[(rc7['CROP']==crop), 'CONFAC'] = ccf
        
    
    conn.close()
    return rc7
#%%



#%%
if __name__ == "__main__":
    
    os.chdir("..")
    os.chdir("02_meta_data")
    
    #Details from Meta_Data file
    state = pd.read_excel('Meta_Data.xlsx', sheet_name='state')
    st = pd.read_excel('Meta_Data.xlsx', sheet_name='st')
    
    #AS1.0 details from Meta_Data file
    as1p0_season = pd.read_excel('Meta_Data.xlsx', sheet_name='as1p0_season')
    as1p0_sch = pd.read_excel('Meta_Data.xlsx', sheet_name='as1p0_sch')
    as1p0_crops = pd.read_excel('Meta_Data.xlsx', sheet_name='as1p0_crops')

    #AS2.0 details from Meta_Data file
    as2p0_season = pd.read_excel('Meta_Data.xlsx', sheet_name='as2p0_season')
    as2p0_sch = pd.read_excel('Meta_Data.xlsx', sheet_name='as2p0_sch')
    as2p0_sch = as2p0_sch.melt(['season_cd','crop_cd','crop_nm','ccf'], var_name='st_cd', value_name='nexp')
    as2p0_sch['st_cd'] = as2p0_sch['st_cd'].str.replace('plan_cen','1').str.replace('plan_state','2').astype(int)

    #setting variables
    state_code = int(state.loc[0,'state_cd'])
    year = int(state.loc[0,'year'])
    st_codes = list(st['st_cd'])
    season_codes_as1p0 = list(as1p0_season['season_cd'])
    season_codes_as2p0 = list(as2p0_season['season_cd'])
    
    #changing directory to parent folder
    os.chdir("..")
    
    #creating blank dataframes
    rc1 = pd.DataFrame()
    rc2 = pd.DataFrame()
    rc3 = pd.DataFrame()
    rc7 = pd.DataFrame()

    #AS1.0
    for season_code in season_codes_as1p0:
        for st_code in st_codes:
            details_1p0 = {
                'year' : year,
                'state_cd' : state_code,
                'season_cd' : season_code,
                'st_cd' : st_code,
                'novp' : int(as1p0_sch[(as1p0_sch['season_cd']==season_code) & (as1p0_sch['st_cd']==st_code)]['novp']),
                'novcod' : int(as1p0_sch[(as1p0_sch['season_cd']==season_code) & (as1p0_sch['st_cd']==st_code)]['novcod']),
                'ddg' : int(as1p0_sch[(as1p0_sch['season_cd']==season_code) & (as1p0_sch['st_cd']==st_code)]['ddg']),
                'ddtrs' : int(as1p0_sch[(as1p0_sch['season_cd']==season_code) & (as1p0_sch['st_cd']==st_code)]['ddtrs']),
                'crop_set' : set(as1p0_crops[as1p0_crops['season_cd']==season_code]['crop_cd'])
            }
            r1, r2, r3 = clean_rc123(details_1p0)

            rc1 = pd.concat([rc1, r1], ignore_index = True)
            rc2 = pd.concat([rc2, r2], ignore_index = True)
            rc3 = pd.concat([rc3, r3], ignore_index = True)


    #AS2.0
    for season_code in season_codes_as2p0:
        for st_code in st_codes:
            details_2p0 = {
                'year' : year,
                'state_cd' : state_code,
                'season_cd' : season_code,
                'st_cd' : st_code,
                'crop_details' : as2p0_sch[(as2p0_sch['st_cd']==1) & (as2p0_sch['season_cd']==1)][['crop_cd','nexp','ccf']]
            }
            r7 = clean_rc7(details_2p0)
            rc7 = pd.concat([rc7, r7], ignore_index = True)
    
    #deleting pre existing log.txt file.
    log_path = 'log.txt'
    if os.path.exists(log_path):
        os.remove(log_path)

    #Check if HSSN is not present
    if (rc1[rc1['HSSN'].isna()].shape[0] + rc1[rc1['HSSN'] == 0].shape[0]) > 0:
        with open("log.txt", "a") as f:
            print('='*80, file=f)
            print('In RC1, HSSN is Zero or Blank. Please Correct...', file=f)
            print('='*80, file=f)

    #Check if GEOA is not present
    if (rc1[rc1['GEOA'].isna()].shape[0] + rc1[rc1['GEOA'] == 0].shape[0]) > 0:
        with open("log.txt", "a") as f:
            print('='*80, file=f)
            print('In RC1, GEOA is Zero or Blank. Please Correct...', file=f)
            print('='*80, file=f)
    
    #RC2 area abnormal changes alert
    with open("log.txt", "a") as f:
        print('='*80, file=f)
        print('In RC2, Check if there is abnormal change in Supervisor & Patwari area...', file=f)
        print('='*80, file=f)

    rc1.sort_values(['ST', 'SESON', 'DIST', 'STRA', 'VILL'], inplace=True)
    rc2.sort_values(['ST', 'SESON', 'DIST', 'STRA', 'VILL', 'CROP', 'VARC'], inplace=True)
    rc3.sort_values(['ST', 'SESON', 'DIST', 'STRA', 'VILL', 'SN', 'CROP'], inplace=True)
    rc7.sort_values(['ST', 'SESON', 'CROP', 'DIST', 'VILL', 'EXPNO', 'IRR'], inplace=True)


    #changing directory to RCs
    os.chdir("04_Clean_RCs")

    #Delete RC1237 if exists from before
    rc1237 = "RC1237_clean.xlsx"
    if os.path.exists(rc1237):
        os.remove(rc1237)

    #Create RC1237 excel file
    wb = openpyxl.Workbook()

    #RC1
    ws1 = wb.create_sheet(title='RC1')
    for r in dataframe_to_rows(rc1, index=False, header=True):
        ws1.append(r)

    #RC2
    ws2 = wb.create_sheet(title='RC2')
    for r in dataframe_to_rows(rc2, index=False, header=True):
        ws2.append(r)

    #RC3
    ws3 = wb.create_sheet(title='RC3')
    for r in dataframe_to_rows(rc3, index=False, header=True):
        ws3.append(r)

    #RC7
    ws7 = wb.create_sheet(title='RC7')
    for r in dataframe_to_rows(rc7, index=False, header=True):
        ws7.append(r)

    #remove unnecessary sheet
    sheet_to_remove = 'Sheet'
    if sheet_to_remove in wb.sheetnames:
        del wb[sheet_to_remove]
    
    wb.save('RC1237_clean_v1.xlsx')

#%%

