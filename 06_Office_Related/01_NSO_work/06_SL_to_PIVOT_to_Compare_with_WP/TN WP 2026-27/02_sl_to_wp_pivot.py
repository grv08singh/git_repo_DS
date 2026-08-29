import os
import glob
import pandas as pd
import xlrd
import re
from pathlib import Path
import numpy as np
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, PatternFill



def format_excel(excel_path):
    
    wb = load_workbook(excel_path)
    #ws = wb.active

    for ws in wb.worksheets:
        wb.active = ws
        # filling colors
        season_colors = {
            'Kharif': 'FFFFFFDD',   # light yellow
            'Rabi': 'FFE7FFFF'      # light cyan
        }
        header_row = 1
        max_col = ws.max_column
        col_season = {}
        current_season = None
    
        for col_idx in range(1, max_col + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value is not None and cell_value in season_colors:
                current_season = cell_value
            # If the cell is blank, it belongs to the previous season (if any)
            if current_season is not None:
                col_season[col_idx] = current_season
    
        for col_idx, season in col_season.items():
            fill = PatternFill(start_color=season_colors[season],
                               end_color=season_colors[season],
                               fill_type='solid')
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                cell.fill = fill
    
        # making border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
    
        # rotate row 2
        rotation_alignment = Alignment(textRotation=90, horizontal='center', vertical='bottom')
        for cell in ws[2]:
            cell.alignment = rotation_alignment
        
        # Set column widths (add a little padding)
        for col_index in range(1, ws.max_column+1):
            col_letter = get_column_letter(col_index)
            if col_index == 2:
                ws.column_dimensions[col_letter].width = 20
            else:
                ws.column_dimensions[col_letter].width = 3.6
    
    wb.save(excel_path)
	
	



def get_SL_to_WP(directory, st):
    records = []

    distt_seq = {
        "Erode": 1,
        "The Nilgiris": 2,
        "Coimbatore": 3,
        "Tiruppur": 4,
        "Dharmapuri": 5,
        "Krishnagiri": 6,
        "Thirupathur": 7,
        "Salem": 8,
        "Namakkal": 9,
        "Karur": 10,
        "Tiruchirappalli": 11,
        "Perambalur": 12,
        "Ariyalur": 13,
        "Thiruvallur": 14,
        "kancheepuram": 15,
        "Chengalpattu": 16,
        "Villupuram": 17,
        "Cuddalore": 18,
        "Kallakuruchi": 19,
        "Mayiladuthurai": 20,
        "Vellore": 21,
        "Tiruvannamalai": 22,
        "Ranipet": 23,
        "Dindigul": 24,
        "Sivagangai": 25,
        "Madurai": 26,
        "Theni": 27,
        "Nagapattinam": 28,
        "Thiruvarur": 29,
        "Thanjavur": 30,
        "Pudukkottai": 31,
        "Thoothukkudi": 32,
        "Thirunelveli": 33,
        "Kanyakumari": 34,
        "Tenkasi": 35,
        "Virudhunagar": 36,
        "Ramanathapuram": 37
    }
    
    # Get all .xlsx and .xls files in the directory
    excel_files = glob.glob(os.path.join(directory, "*.xlsx"))
    df = pd.DataFrame()

    # Global Variables
    kh = 'Kharif'
    rb = 'Rabi'
    kh_keywords1 = ('kar','kur','sam')
    rb_keywords1 = ('nav','rai')
    
    kh_keywords2 = ('K','sugar','can','maiz','ze','red','dgram')
    rb_keywords2 = ('R','®','hor','rse','segram','hour','green','engram','kgram','black')
    
    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        distt = Path(file_name).stem
        # Determine file type and use appropriate library
        if file_path.endswith('.xlsx'):
            try:
                wb = load_workbook(file_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    if sheet_name.lower() in st:
                        ws = wb[sheet_name]
                        
                        # setting which cells to scan
                        start_row = 6
                        start_col = 2
                        end_row = ws.max_row + 1
                        end_col = ws.max_column + 1

                        for col_idx in range(start_col, end_col):
                            # get crop name and type
                            crop = ws.cell(row=4, column=col_idx).value
                            if isinstance(crop, str) and crop is not None:
                                crop = crop.strip().replace(' ','')
                            else:
                                crop = ws.cell(row=4, column=col_idx-1).value
                                if isinstance(crop, str) and crop is not None:
                                    crop = crop.strip().replace(' ','')
                                else:
                                    crop = ws.cell(row=4, column=col_idx-2).value
                                    if isinstance(crop, str) and crop is not None:
                                        crop = crop.strip().replace(' ','')
                                    else:
                                        crop = None

                            # set season
                            if crop is not None:
                                crop = re.sub(r'\s+', '', crop)
                                crop = re.sub(r'UI', '', crop)
                                
                                crop_type = ws.cell(row=5, column=col_idx).value
                                if isinstance(crop_type, str) and crop_type is not None:
                                    crop_type = crop_type.strip()

                                if crop_type is not None and any(k in crop_type.lower() for k in kh_keywords1):
                                    season = kh
                                elif crop_type is not None and any(r in crop_type.lower() for r in rb_keywords1):
                                    season = rb
                                elif 'K' in crop or any(k in crop.lower() for k in kh_keywords2):
                                    season = kh
                                elif 'R' in crop or any(r in crop.lower() for r in rb_keywords2):
                                    season = rb
                                else:
                                    season = '----'
    
                                if crop_type == 'I':
                                    crop_type = 'IR'

                                # set crop names correctly
                                crop1 = ''
                                if 'pad' in crop.lower():
                                    if 'kar' in crop_type.lower():
                                        crop1 = 'paddy_kar'
                                    if 'sam' in crop_type.lower():
                                        crop1 = 'paddy_samba'
                                    if any(nvr in crop_type.lower() for nvr in rb_keywords1):
                                        crop1 = 'paddy_navrai'
                                
                                if any(bjr in crop.lower() for bjr in ('baj','jra','jara')):
                                    crop1 = 'bajra'
                                if any(ctn in crop.lower() for ctn in ('cot','ton')):
                                    crop1 = 'cotton'
                                if any(ggl in crop.lower() for ggl in ('gin','gel')):
                                    crop1 = 'gingelly'
                                if any(gnut in crop.lower() for gnut in ('ground','ndnut')):
                                    crop1 = 'groundnut'
                                if any(mze in crop.lower() for mze in ('maiz','ze')):
                                    crop1 = 'maize'
                                if any(bram in crop.lower() for bram in ('black','kgram')):
                                    crop1 = 'blackgram'
                                if any(sc in crop.lower() for sc in ('sugar','can')):
                                    crop1, crop_type = 'sugarcane', '--'
                                if any(jwar in crop.lower() for jwar in ('jow','war')):
                                    crop1 = 'jowar'
                                if any(hgrm in crop.lower() for hgrm in ('hor','rse','segram','hour')):
                                    crop1 = 'horsegram'
                                if any(ggm in crop.lower() for ggm in ('green','engram')):
                                    crop1 = 'greengram'
                                if any(rgm in crop.lower() for rgm in ('red','dgram')):
                                    crop1 = 'redgram'

                                if crop1 == 'jowar':
                                    temp = crop.lower().replace('jowar','')
                                    if 'k' in temp:
                                        season = kh
                                    elif any(r in temp for r in ('r','®')):
                                        season = rb
                                    else:
                                        season = '--'

                                if 'paddy' in crop1:
                                    crop_type = '--'

                                vill_count = 0
                                for row_idx in range(start_row, end_row):
                                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                                    if isinstance(cell_val, str) and cell_val is not None:
                                        cell_val = cell_val.strip()
    
                                    if cell_val is not None:
                                        if '-' in cell_val:
                                            if re.search(r'\d+-\d+', cell_val):
                                                vill_count += 1
                                plan = 2 * vill_count
                                if plan != 0:
                                    new_row = {
                                        #'CROP_orig': crop,
                                        'CROP': crop1,
                                        'CT': crop_type,
                                        'SEASON': season,
                                        'DISTT': distt,
                                        'PLAN': plan,
                                        'SEQ': distt_seq[distt]
                                    }
                                    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                                    df = df.sort_values(['SEQ','SEASON','CROP','CT'])
                                    df_pivot = df.pivot_table(index=['SEQ','DISTT'],columns=['SEASON','CROP','CT'],values='PLAN',fill_value=np.nan,aggfunc='sum')
                                    df_pivot = df_pivot.sort_index()
                                    df_pivot = df_pivot.sort_index(axis=1, level=[0, 1, 2])

            except Exception as e:
                print(f"Error reading {file_name}, {st[0]}, {crop}, {crop_type}, {season}, {distt}, {plan} (openpyxl): {e}")
    return df_pivot





if __name__ == "__main__":
    curr_dir = Path.cwd()
    directory = curr_dir / "SL 2.0"
    excel_path = curr_dir / "SL_to_WP.xlsx"
    
    with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
        for sample in range(1,3):
            if sample == 1:
                st = ['central', 'centre', 'center']
            elif sample == 2:
                st = ['state', 'statee']
            
            df_pivot = get_SL_to_WP(directory, st)
            df_pivot.to_excel(writer, sheet_name=f"{st[0]}")

    # format excel workbook
    format_excel(excel_path)