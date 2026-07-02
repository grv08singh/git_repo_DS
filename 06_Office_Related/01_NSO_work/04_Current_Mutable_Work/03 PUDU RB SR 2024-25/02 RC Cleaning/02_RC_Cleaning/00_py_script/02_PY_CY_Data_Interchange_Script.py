import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np



def handle_HSSN_GEOA():
    
    rc1 = pd.read_excel("RC1237_clean_v1.xlsx", sheet_name="RC1")
    rc2 = pd.read_excel("RC1237_clean_v1.xlsx", sheet_name="RC2")
    rc3 = pd.read_excel("RC1237_clean_v1.xlsx", sheet_name="RC3")
    rc7 = pd.read_excel("RC1237_clean_v1.xlsx", sheet_name="RC7")
    
    missing_hssn_count = ((rc1[rc1['HSSN'].isna()].shape[0]) + (rc1[rc1['HSSN']==0].shape[0]))
    missing_geoa_count = ((rc1[rc1['GEOA'].isna()].shape[0]) + (rc1[rc1['GEOA']==0].shape[0]))
    
    if missing_hssn_count > 0:
        median_HSSN = int(rc1['HSSN'].median())
        rc1['HSSN'] = rc1['HSSN'].fillna(median_HSSN)
        rc1['HSSN'] = rc1['HSSN'].replace(0, median_HSSN)
        rc2['HSSN'] = rc2['HSSN'].fillna(median_HSSN)
        rc2['HSSN'] = rc2['HSSN'].replace(0, median_HSSN)
    
    if missing_geoa_count > 0:
        mean_GEOA = np.round(rc1['GEOA'].mean(), 2)
        rc1['GEOA'] = rc1['GEOA'].fillna(mean_GEOA)
        rc1['GEOA'] = rc1['GEOA'].replace(0,mean_GEOA)
        
    #Create RC1237_v2 excel file
    wb = openpyxl.Workbook()

    #RC1
    ws1 = wb.create_sheet(title='RC1')
    for r in dataframe_to_rows(rc1, index=False, header=True):
        ws1.append(r)

    #RC2
    ws2 = wb.create_sheet(title='RC2')
    for r in dataframe_to_rows(rc2, index=False, header=True):
        ws2.append(r)

    #RC3
    ws3 = wb.create_sheet(title='RC3')
    for r in dataframe_to_rows(rc3, index=False, header=True):
        ws3.append(r)

    #RC7
    ws7 = wb.create_sheet(title='RC7')
    for r in dataframe_to_rows(rc7, index=False, header=True):
        ws7.append(r)
    
    #remove unnecessary sheet
    sheet_to_remove = 'Sheet'
    if sheet_to_remove in wb.sheetnames:
        del wb[sheet_to_remove]
    
    wb.save('RC1237_clean_v2.xlsx')
    


def empty_py(wb_py):
    #deleting data from RC1 PY
    ws_py1 = wb_py["RC-1(PY)"]
    max_row = ws_py1.max_row
    ws_py1.delete_rows(1, max_row)
    
    #deleting data from RC2 PY
    ws_py2 = wb_py["RC-2(PY)"]
    max_row = ws_py2.max_row
    ws_py2.delete_rows(1, max_row)
    
    #deleting data from RC3 PY
    ws_py3 = wb_py["RC-3(PY)"]
    max_row = ws_py3.max_row
    ws_py3.delete_rows(1, max_row)
    
    #deleting data from RC7 PY
    ws_py7 = wb_py["RC-7(PY)"]
    max_row = ws_py7.max_row
    ws_py7.delete_rows(1, max_row)
    
    return wb_py


def fill_py(wb_py, wb_cy):
    
    #copying RC1 CY to PY
    ws_py1 = wb_py["RC-1(PY)"]
    ws_cy1 = wb_cy["RC-1"]
    for row in ws_cy1.iter_rows():
        for cell in row:
            ws_py1[cell.coordinate].value = cell.value
    
    #copying RC2 CY to PY
    ws_py2 = wb_py["RC-2(PY)"]
    ws_cy2 = wb_cy["RC-2"]
    for row in ws_cy2.iter_rows():
        for cell in row:
            ws_py2[cell.coordinate].value = cell.value
    
    #copying RC3 CY to PY
    ws_py3 = wb_py["RC-3(PY)"]
    ws_cy3 = wb_cy["RC-3"]
    for row in ws_cy3.iter_rows():
        for cell in row:
            ws_py3[cell.coordinate].value = cell.value
    
    #copying RC7 CY to PY
    ws_py7 = wb_py["RC-7(PY)"]
    ws_cy7 = wb_cy["RC-7"]
    for row in ws_cy7.iter_rows():
        for cell in row:
            ws_py7[cell.coordinate].value = cell.value
    
    return wb_py


def empty_cy(wb_cy):

    #deleting data from RC1 CY
    ws_cy1 = wb_cy["RC-1"]
    max_row = ws_cy1.max_row
    ws_cy1.delete_rows(1, max_row)
    
    #deleting data from RC2 CY
    ws_cy2 = wb_cy["RC-2"]
    max_row = ws_cy2.max_row
    ws_cy2.delete_rows(1, max_row)
    
    #deleting data from RC3 CY
    ws_cy3 = wb_cy["RC-3"]
    max_row = ws_cy3.max_row
    ws_cy3.delete_rows(1, max_row)
    
    #deleting data from RC7 CY
    ws_cy7 = wb_cy["RC-7"]
    max_row = ws_cy7.max_row
    ws_cy7.delete_rows(1, max_row)
    
    return wb_cy


def fill_cy(wb_cy, wb_rc1237):
    
    #copying newly created RC1 to CY
    ws_cy1 = wb_cy["RC-1"]
    ws_rc1 = wb_rc1237["RC1"]
    for row in ws_rc1.iter_rows():
        for cell in row:
            ws_cy1[cell.coordinate].value = cell.value

    #copying newly created RC2 to CY
    ws_cy2 = wb_cy["RC-2"]
    ws_rc2 = wb_rc1237["RC2"]
    for row in ws_rc2.iter_rows():
        for cell in row:
            ws_cy2[cell.coordinate].value = cell.value

    #copying newly created RC3 to CY
    ws_cy3 = wb_cy["RC-3"]
    ws_rc3 = wb_rc1237["RC3"]
    for row in ws_rc3.iter_rows():
        for cell in row:
            ws_cy3[cell.coordinate].value = cell.value

    #copying newly created RC7 to CY
    ws_cy7 = wb_cy["RC-7"]
    ws_rc7 = wb_rc1237["RC7"]
    for row in ws_rc7.iter_rows():
        for cell in row:
            ws_cy7[cell.coordinate].value = cell.value
    
    return wb_cy




if __name__ == "__main__":
    
    #handle missing/zero HSSN, GEOA
    os.chdir('..')
    os.chdir('04_Clean_RCs')
    handle_HSSN_GEOA()
    
    #loading newly created RCs
    wb_rc1237 = openpyxl.load_workbook("RC1237_clean_v2.xlsx")
    ws_rc1 = wb_rc1237["RC1"]
    ws_rc2 = wb_rc1237["RC2"]
    ws_rc3 = wb_rc1237["RC3"]
    ws_rc7 = wb_rc1237["RC7"]

    #read previous year's CY & PY
    os.chdir('..')
    os.chdir('03_PY_CY_Old')
    wb_py = openpyxl.load_workbook("PY State Details & Capital-N .xlsx")
    wb_cy = openpyxl.load_workbook("CY State Details & Capital-N.xlsx")

    #updating CY, PY files
    wb_py = empty_py(wb_py)
    wb_py = fill_py(wb_py, wb_cy)
    wb_cy = empty_cy(wb_cy)
    wb_cy = fill_cy(wb_cy, wb_rc1237)

    #saving new PY, CY excel files in folder "06_PY_CY_New"
    os.chdir('..')
    os.chdir('05_PY_CY_New')
    wb_py.save("PY State Details & Capital-N .xlsx")
    wb_cy.save("CY State Details & Capital-N.xlsx")